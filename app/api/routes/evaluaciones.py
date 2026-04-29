from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
import uuid

from app.models.pregunta import Pregunta
from app.models.respuesta import Respuesta


from app.services.calculo_nom035 import calcular_resultados
from app.services.reporte_ejecutivo_nom035 import generar_reporte_ejecutivo
from app.services.interpretacion_nom035 import interpretar_resultados
from app.api.deps import get_db, get_current_user
from app.models.evaluacion import Evaluacion
from app.models.asignacion import Asignacion
from app.models.cuestionario import Cuestionario
from app.models.usuario import Usuario
from app.models.empresa import Empresa  # 🔥 IMPORT CORRECTO
from app.models.orden import Orden
from app.schemas.evaluacion import GenerarEvaluacionSchema
from openpyxl import load_workbook
import os
from fastapi.responses import FileResponse
import tempfile
from app.core.config import settings

router = APIRouter()


# =========================
# 🧠 LÓGICA NOM-035
# =========================
def obtener_tipo_y_escala(total_empleados: int):
    if total_empleados <= 15:
        return "I", "BINARIA"
    elif total_empleados <= 50:
        return "II", "LIKERT"
    else:
        return "III", "LIKERT"


# =========================
# 🔥 GENERAR EVALUACIÓN
# =========================
@router.post("/empresa/evaluaciones")
def generar_evaluacion(
    data: GenerarEvaluacionSchema,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user)
):

    empresa_id = user.empresa_id
    total = data.total_empleados

    # =========================
    # 🧠 VALIDACIÓN BASE
    # =========================
    if total <= 0:
        raise HTTPException(400, "Total de empleados inválido")

    # =========================
    # 🔒 VALIDAR LÍMITE POR PLAN (SaaS real)
    # =========================
    empresa = db.query(Empresa).filter_by(id=empresa_id).first()

    if empresa and empresa.plan:

        limites = {
            "micro": 15,
            "mediana": 50,
            "grande": 250,
            "corporativo": 999999
        }

        limite = limites.get(empresa.plan)

        if limite is None:
            raise HTTPException(500, "Plan inválido configurado")

        if total > limite:
            raise HTTPException(
                status_code=403,
                detail=f"Tu plan {empresa.plan} permite máximo {limite} empleados. Actualiza tu plan para continuar."
            )

    # =========================
    # 🔒 VALIDAR CENTROS (lógica negocio)
    # =========================
    orden_principal = db.query(Orden)\
        .filter_by(
            user_id=user.id,
            tipo="principal",
            estado="pagado"
        )\
        .first()

    if orden_principal:
        evaluaciones_pagadas = db.query(Evaluacion)\
            .filter_by(
                empresa_id=empresa_id,
                pagado=True
            ).count()

        if evaluaciones_pagadas >= 4:
            raise HTTPException(
                status_code=403,
                detail="Has alcanzado el límite de centros de trabajo."
            )

    # =========================
    # 🧠 CONFIGURACIÓN NOM-035
    # =========================
    tipo, tipo_escala = obtener_tipo_y_escala(total)

    cuestionario = db.query(Cuestionario).filter_by(tipo=tipo).first()

    if not cuestionario:
        raise HTTPException(404, f"No existe cuestionario tipo {tipo}")

    # =========================
    # 🚫 EVITAR DUPLICADOS
    # =========================
    evaluacion_existente = db.query(Evaluacion).filter_by(
        empresa_id=empresa_id,
        pagado=False
    ).first()

    if evaluacion_existente:
        print("⚠️ Reutilizando evaluación pendiente")

        asignaciones = db.query(Asignacion).filter_by(
            evaluacion_id=evaluacion_existente.id
        ).all()

        tokens = [{
            "token": a.token,
            "link": f"{settings.FRONTEND_URL}/cuestionario/{a.token}"
        } for a in asignaciones]

        return {
            "id": str(evaluacion_existente.id),
            "tipo": tipo,
            "tipo_escala": evaluacion_existente.tipo_escala,
            "total_empleados": len(asignaciones),
            "tokens": tokens
        }

    # =========================
    # 🧾 CREAR EVALUACIÓN
    # =========================
    evaluacion = Evaluacion(
        empresa_id=empresa_id,
        cuestionario_id=cuestionario.id,
        tipo_escala=tipo_escala,
        pagado=False
    )

    db.add(evaluacion)
    db.flush()

    tokens = []

    for _ in range(total):
        token = str(uuid.uuid4())

        asignacion = Asignacion(
            evaluacion_id=evaluacion.id,
            cuestionario_id=cuestionario.id,
            token=token,
            completado=False
        )

        db.add(asignacion)

        tokens.append({
            "token": token,
            "link": f"{settings.FRONTEND_URL}/cuestionario/{token}"
        })

    db.commit()
    db.refresh(evaluacion)

    return {
        "id": str(evaluacion.id),
        "tipo": tipo,
        "tipo_escala": tipo_escala,
        "total_empleados": total,
        "tokens": tokens
    }


# =========================
# 🔥 LISTAR EVALUACIONES
# =========================
@router.get("/empresa/evaluaciones")
def listar_evaluaciones(
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user)
):

    evaluaciones = db.query(Evaluacion)\
        .filter_by(empresa_id=user.empresa_id)\
        .order_by(Evaluacion.id.desc())\
        .all()

    resultado = []

    for e in evaluaciones:

        asignaciones = db.query(Asignacion).filter_by(
            evaluacion_id=e.id
        ).all()

        total = len(asignaciones)
        respondidos = sum(1 for a in asignaciones if a.completado)

        resultado.append({
            "id": str(e.id),
            "tipo_escala": e.tipo_escala,
            "total_empleados": total,
            "respondidos": respondidos,
            "progreso": round((respondidos / total) * 100) if total > 0 else 0,
            "pagado": e.pagado  # 👈 CLAVE
        })

    return resultado


# =========================
# 🔥 PENDIENTES
# =========================
@router.get("/empresa/evaluacion/{evaluacion_id}/pendientes")
def obtener_pendientes(
    evaluacion_id: str,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user)
):

    evaluacion = db.query(Evaluacion).filter_by(
        id=evaluacion_id,
        empresa_id=user.empresa_id
    ).first()

    if not evaluacion:
        raise HTTPException(404, "Evaluación no encontrada")

    asignaciones = db.query(Asignacion).filter_by(
        evaluacion_id=evaluacion_id
    ).all()

    return [
        {
            "token": a.token,
            "status": "Respondido" if a.completado else "Pendiente",
            "link": f"{settings.FRONTEND_URL}/cuestionario/{a.token}"
        }
        for a in asignaciones
    ]


# =========================
# 📊 RESULTADOS PRO (FIX REAL)
# =========================
@router.get("/empresa/resultados/{evaluacion_id}")
def obtener_resultados(
    evaluacion_id: str,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user)
):

    # =========================
    # 🔒 VALIDAR EVALUACIÓN
    # =========================
    evaluacion = db.query(Evaluacion).filter_by(
        id=evaluacion_id,
        empresa_id=user.empresa_id
    ).first()

    if not evaluacion:
        raise HTTPException(404, "Evaluación no encontrada")

    if not evaluacion.pagado:
        raise HTTPException(402, "Pago requerido")

    # 🔥 OBTENER TIPO REAL (CLAVE)
    cuestionario = db.query(Cuestionario).filter_by(
        id=evaluacion.cuestionario_id
    ).first()

    tipo_cuestionario = cuestionario.tipo
    tipo_escala = evaluacion.tipo_escala

    # =========================
    # 📊 RESPUESTAS
    # =========================
    respuestas = db.query(Respuesta)\
        .join(Asignacion)\
        .join(Pregunta)\
        .filter(Asignacion.evaluacion_id == evaluacion_id)\
        .all()

    if not respuestas:
        return {
            "global": {"puntaje": 0, "nivel": "Sin datos"},
            "dominios": {},
            "interpretacion": {},
            "reporte_ejecutivo": {
                "resumen": "Sin respuestas suficientes",
                "analisis_dominios": [],
                "plan": [],
                "dominio_critico": None
            },
            "total_empleados": 0,
            "respondidos": 0,
            "progreso": 0
        }

    # =========================
    # 📚 PREGUNTAS
    # =========================
    preguntas = db.query(Pregunta).all()
    preguntas_dict = {p.id: p for p in preguntas}

    # =========================
    # 🔥 CÁLCULO CORRECTO
    # =========================
    resultados = calcular_resultados(
        respuestas,
        preguntas_dict,
        tipo_cuestionario=tipo_cuestionario,
        tipo_escala=tipo_escala
    )

    # =========================
    # 👥 DATOS REALES
    # =========================
    asignaciones = db.query(Asignacion).filter_by(
        evaluacion_id=evaluacion_id
    ).all()

    total_empleados = len(asignaciones)
    respondidos = sum(1 for a in asignaciones if a.completado)

    # =========================
    # 🧠 INTERPRETACIÓN
    # =========================
    interpretacion = interpretar_resultados(resultados)

    # =========================
    # 📄 REPORTE EJECUTIVO
    # =========================
    reporte = generar_reporte_ejecutivo(
        empresa=str(user.empresa_id),
        empleados=total_empleados,
        resultados=resultados,
        interpretacion=interpretacion
    )

    # =========================
    # 🚀 RESPUESTA FINAL
    # =========================
    return {
        "global": resultados["global"],
        "dominios": resultados["dominios"],
        "interpretacion": interpretacion,
        "reporte_ejecutivo": reporte,
        "total_empleados": total_empleados,
        "respondidos": respondidos,
        "progreso": round(
            (respondidos / total_empleados) * 100
        ) if total_empleados > 0 else 0
    }


# =========================
# 📊 EXPORTAR EXCEL (.xlsm con macros)
# =========================
@router.get("/empresa/evaluacion/{evaluacion_id}/excel")
def exportar_excel(
    evaluacion_id: str,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user)
):
    try:

        # =========================
        # 🔒 VALIDACIONES
        # =========================
        evaluacion = db.query(Evaluacion).filter_by(
            id=evaluacion_id,
            empresa_id=user.empresa_id
        ).first()

        if not evaluacion:
            raise HTTPException(403, "No tienes acceso")

        if not evaluacion.pagado:
            raise HTTPException(403, "Debes pagar")

        asignaciones = db.query(Asignacion).filter_by(
            evaluacion_id=evaluacion_id
        ).all()

        # =========================
        # 📂 BUSCAR PLANTILLA (ROBUSTO)
        # =========================
        posibles_rutas = [
            os.path.join(os.getcwd(), "plantilla_nom035.xlsm"),
            os.path.join(os.getcwd(), "app", "plantilla_nom035.xlsm"),
            os.path.join(os.path.dirname(__file__), "..", "plantilla_nom035.xlsm"),
        ]

        ruta_plantilla = None

        for ruta in posibles_rutas:
            ruta_abs = os.path.abspath(ruta)
            print("🔍 Probando ruta:", ruta_abs)
            if os.path.exists(ruta_abs):
                ruta_plantilla = ruta_abs
                break

        if not ruta_plantilla:
            raise HTTPException(500, "No se encontró plantilla_nom035.xlsm en ninguna ruta")

        print("✅ Usando plantilla:", ruta_plantilla)

        # =========================
        # 📄 CARGAR EXCEL CON MACROS
        # =========================
        wb = load_workbook(
            ruta_plantilla,
            keep_vba=True,
            data_only=False
        )

        ws = wb.active

        # =========================
        # 🧹 LIMPIAR
        # =========================
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)

        # =========================
        # ✍️ LLENAR DATOS
        # =========================
        fila = 2

        for a in asignaciones:
            ws.cell(row=fila, column=1, value=a.token)
            ws.cell(row=fila, column=2, value="Respondido" if a.completado else "Pendiente")
            ws.cell(row=fila, column=3, value=f"{settings.FRONTEND_URL}/cuestionario/{a.token}")
            ws.cell(row=fila, column=4, value="")
            ws.cell(row=fila, column=5, value="")
            fila += 1

        # =========================
        # 💾 GUARDAR
        # =========================
        temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsm")
        temp.close()

        wb.save(temp.name)

        print("📁 Archivo generado:", temp.name)

        # =========================
        # 📤 RESPUESTA
        # =========================
        return FileResponse(
            path=temp.name,
            filename="evaluacion_nom035.xlsm",
            media_type="application/vnd.ms-excel.sheet.macroEnabled.12"
        )

    except Exception as e:
        print("💥 ERROR REAL:", str(e))
        raise HTTPException(500, f"Error generando Excel: {str(e)}")